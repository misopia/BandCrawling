

import pandas as pd
import os
import math
from datetime import datetime 


from openpyxl import Workbook


dir_path = "D:/Project/GIT/BandCrawling/BandCrawling/data/"

lstXlsPath = []

for (root, directories, files) in os.walk(dir_path):
    for d in directories:
        d_path = os.path.join(root, d)
        #print(d_path)

    for file in files:
        file_path = os.path.join(root, file)
        #print(file_path)
        if file_path.find(".zip") < 0 and file_path.find(".lnk") < 0 and file_path.find(".pdf") < 0:
            lstXlsPath.append(file_path)

#print(lstXlsPath)


#엑셀 데이터 읽어오기
lstExcelAllData = []
for XlsPath in lstXlsPath : 
    df = pd.read_excel(XlsPath)
    print(XlsPath)
    #입고일자
    strInDate = ""
    lstData = []

    for index, row in df.iterrows():
        
        #print(row, end = "\n\n")

        if index == 1 :
            strInDate = row[2]
        elif index >= 17 :
            if math.isnan(row[1]) == True :
                break

            lstData.append(row)

    lstExcelAllData.append([strInDate,lstData])
    #print(len(lstExcelAllData))
    
    #if len(lstExcelAllData) == 100 :
    #    break
    #print(lstData)



write_wb = Workbook()
#write_wb.save('D:/Project/GIT/BandCrawling/BandCrawling/MergeTTT{0}.xlsx'.format(datetime.now().strftime('%Y-%m-%d_%H%M%S')))

write_ws = write_wb.active


idxRow = 2
for excelAllData in lstExcelAllData :
    for dataTT in excelAllData[1] :
        idxCol = 1
        idxT = 0
        for dataT in dataTT :
            if idxT == 0 :
                idxT +=1
                continue

            idxT +=1
            #if write_ws.cell(idxRow,idxCol) == "test" :
            #    abbbb = 0

            write_ws.cell(idxRow,idxCol, dataT)
            idxCol += 1
            #if dataT == 'test' :
            #    abbbb = 0
    
        write_ws.cell(idxRow,idxCol, excelAllData[0])
        idxRow += 1

            







write_wb.save('D:/Project/GIT/BandCrawling/BandCrawling/Merge{0}.xlsx'.format(datetime.now().strftime('%Y-%m-%d_%H%M%S')))



